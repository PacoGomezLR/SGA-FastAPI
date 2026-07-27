from datetime import datetime
from io import BytesIO

from fastapi import HTTPException, status
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.repositories.stock_repository import StockRepository
from app.repositories.product_repository import ProductRepository
from app.repositories.location_repository import LocationRepository
from app.schemas.stock import StockCreate, StockUpdate


class StockService:
    def __init__(self, db: Session):
        self.db = db
        self.repository = StockRepository(db)
        self.product_repository = ProductRepository(db)
        self.location_repository = LocationRepository(db)

    def get_all_stock(self):
        return self.repository.get_all()

    def get_all_stock_detailed(self, low: bool = False, seccion_id: int | None = None):
        data = self.repository.get_all_with_details(low=low, seccion_id=seccion_id)

        return [
            {
                "id": item.id,
                "producto_id": item.producto_id,
                "producto_nombre": item.producto_nombre,
                "categoria_nombre": item.categoria_nombre,
                "ubicacion_id": item.ubicacion_id,
                "ubicacion_nombre": item.ubicacion_nombre,
                "seccion_id": item.seccion_id,
                "seccion_nombre": item.seccion_nombre,

                "cantidad": item.cantidad,
                "stock_minimo": item.stock_minimo,
                "bajo_stock": item.bajo_stock
            }
            for item in data
        ]

    def export_stock_excel(self, low: bool = False, seccion_id: int | None = None) -> BytesIO:
        filas = self.get_all_stock_detailed(low=low, seccion_id=seccion_id)

        wb = Workbook()
        ws = wb.active
        ws.title = "Stock actual"

        columnas = [
            ("Producto", "producto_nombre"),
            ("Categoría", "categoria_nombre"),
            ("Sección", "seccion_nombre"),
            ("Ubicación", "ubicacion_nombre"),
            ("Cantidad", "cantidad"),
            ("Stock mínimo", "stock_minimo"),
            ("Stock bajo", "bajo_stock"),
        ]

        encabezado_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        encabezado_font = Font(color="FFFFFF", bold=True)
        fila_bajo_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

        for col_index, (titulo, _clave) in enumerate(columnas, start=1):
            celda = ws.cell(row=1, column=col_index, value=titulo)
            celda.fill = encabezado_fill
            celda.font = encabezado_font
            celda.alignment = Alignment(horizontal="left")

        for fila_index, item in enumerate(filas, start=2):
            for col_index, (_titulo, clave) in enumerate(columnas, start=1):
                valor = item.get(clave)
                if clave == "bajo_stock":
                    valor = "Sí" if valor else "No"
                ws.cell(row=fila_index, column=col_index, value=valor if valor is not None else "-")

            if item.get("bajo_stock"):
                for col_index in range(1, len(columnas) + 1):
                    ws.cell(row=fila_index, column=col_index).fill = fila_bajo_fill

        ws.freeze_panes = "A2"

        for col_index, (titulo, clave) in enumerate(columnas, start=1):
            longitud_max = max(
                [len(titulo)] + [len(str(item.get(clave, ""))) for item in filas]
            )
            ws.column_dimensions[get_column_letter(col_index)].width = min(longitud_max + 4, 45)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def get_stock_by_id(self, stock_id: int):
        stock = self.repository.get_by_id(stock_id)
        if not stock:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Registro de stock no encontrado"
            )
        return stock

    def get_stock_by_product(self, producto_id: int):
        return self.repository.get_by_product(producto_id)

    def get_stock_by_location(self, ubicacion_id: int):
        return self.repository.get_by_location(ubicacion_id)

    def create_stock(self, stock_data: StockCreate):
        self.validar_producto_y_ubicacion_existentes(
            stock_data.producto_id,
            stock_data.ubicacion_id
        )

        existing = self.repository.get_by_product_and_location(
            stock_data.producto_id,
            stock_data.ubicacion_id
        )

        if existing:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Ya existe stock para ese producto en esa ubicación"
            )

        return self.repository.create(stock_data)

    def update_stock(self, stock_id: int, stock_data: StockUpdate):
        stock = self.repository.get_by_id(stock_id)

        if not stock:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Registro de stock no encontrado"
            )

        nuevo_producto_id = (
            stock_data.producto_id
            if stock_data.producto_id is not None
            else stock.producto_id
        )

        nueva_ubicacion_id = (
            stock_data.ubicacion_id
            if stock_data.ubicacion_id is not None
            else stock.ubicacion_id
        )

        self.validar_producto_y_ubicacion_existentes(
            nuevo_producto_id,
            nueva_ubicacion_id
        )

        existing = self.repository.get_by_product_and_location(
            nuevo_producto_id,
            nueva_ubicacion_id
        )

        if existing and existing.id != stock_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Ya existe otro registro de stock para ese producto en esa ubicación"
            )

        return self.repository.update(stock, stock_data)

    def delete_stock(self, stock_id: int):
        stock = self.repository.get_by_id(stock_id)

        if not stock:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Registro de stock no encontrado"
            )

        self.repository.delete(stock)

        return {
            "message": "Registro de stock eliminado correctamente"
        }

    def validar_producto_y_ubicacion_existentes(self, producto_id: int, ubicacion_id: int):
        producto = self.product_repository.get_by_id(producto_id)

        if not producto:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="El producto indicado no existe"
            )

        ubicacion = self.location_repository.get_by_id(ubicacion_id)

        if not ubicacion:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="La ubicación indicada no existe"
            )

    def validar_stock_suficiente(self, producto_id: int, ubicacion_id: int, cantidad: int):
        if cantidad <= 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="La cantidad debe ser mayor que cero"
            )

        stock = self.repository.get_by_product_and_location(producto_id, ubicacion_id)

        if not stock:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="No existe stock para ese producto en esa ubicación"
            )

        if stock.cantidad < cantidad:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Stock insuficiente en la ubicación indicada"
            )

        return stock

    def sumar_stock(self, producto_id: int, ubicacion_id: int, cantidad: int):
        if cantidad <= 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="La cantidad a sumar debe ser mayor que cero"
            )

        stock = self.repository.get_by_product_and_location(producto_id, ubicacion_id)

        if stock:
            stock.cantidad += cantidad
            self.db.commit()
            self.db.refresh(stock)
            return stock

        nuevo_stock = StockCreate(
            producto_id=producto_id,
            ubicacion_id=ubicacion_id,
            cantidad=cantidad
        )

        return self.repository.create(nuevo_stock)

    def restar_stock(self, producto_id: int, ubicacion_id: int, cantidad: int):
        if cantidad <= 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="La cantidad a restar debe ser mayor que cero"
            )

        stock = self.validar_stock_suficiente(producto_id, ubicacion_id, cantidad)

        stock.cantidad -= cantidad
        self.db.commit()
        self.db.refresh(stock)

        return stock